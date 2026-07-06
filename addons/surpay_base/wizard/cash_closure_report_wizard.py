from odoo import _, api, fields, models
from odoo.exceptions import UserError
import base64
import io


class SurpayCashClosureReportWizard(models.TransientModel):
    _name = "surpay.cash.closure.report.wizard"
    _description = "Wizard reporte de cierre de caja"

    report_date = fields.Date(
        string="Fecha de cierre",
        required=True,
        default=fields.Date.context_today,
    )
    user_ids = fields.Many2many(
        "res.users",
        string="Usuarios",
        domain=[
            ("active", "=", True),
            ("share", "=", False),
            ("groups_id.category_id.name", "=", "Surpay"),
        ],
    )
    all_surpay_users = fields.Boolean(
        string="Todos los usuarios Surpay",
        default=False,
    )
    can_select_users = fields.Boolean(
        string="Puede seleccionar usuarios",
        default=lambda self: self._current_user_can_select_users(),
        readonly=True,
    )
    current_user_name = fields.Char(
        string="Usuario",
        default=lambda self: self.env.user.name,
        readonly=True,
    )

    @api.model
    def default_get(self, fields_list):
        values = super().default_get(fields_list)
        if "can_select_users" in fields_list:
            values["can_select_users"] = self._current_user_can_select_users()
        if "current_user_name" in fields_list:
            values["current_user_name"] = self.env.user.name
        if "report_date" in fields_list and self.env.context.get("default_report_date"):
            values["report_date"] = self.env.context["default_report_date"]
        if "user_ids" in fields_list and self.env.context.get("default_user_ids"):
            default_user_ids = self.env.context["default_user_ids"]
            if isinstance(default_user_ids, int):
                default_user_ids = [default_user_ids]
            values["user_ids"] = [(6, 0, default_user_ids)]
        if "user_ids" in fields_list and not self._current_user_can_select_users():
            values["user_ids"] = [(6, 0, [self.env.user.id])]
        return values

    @api.model
    def _current_user_can_select_users(self):
        return self.env.user.has_group("base.group_system") or self.env.user.has_group(
            "surpay_base.group_surpay_manager"
        )

    def _get_all_surpay_users(self):
        category = self.env["ir.module.category"].sudo().search(
            [("name", "=", "Surpay")], limit=1
        )
        if not category:
            return self.env["res.users"]
        return self.env["res.users"].sudo().search(
            [
                ("active", "=", True),
                ("share", "=", False),
                ("groups_id.category_id", "=", category.id),
            ]
        )

    def _get_effective_users(self):
        self.ensure_one()
        if self.can_select_users:
            if self.all_surpay_users:
                users = self._get_all_surpay_users()
                if not users:
                    raise UserError(_("No se encontraron usuarios con acceso a Surpay."))
                return users
            if not self.user_ids:
                raise UserError(_("Debes seleccionar al menos un usuario."))
            return self.user_ids
        return self.env.user

    def _report_service(self):
        return self.env["surpay.cash.closure.report.service"]

    def _build_sections_sql(self, users):
        self.ensure_one()
        return self._report_service()._build_cash_closure_sections(
            report_date=self.report_date,
            users=users,
            group_mode="user",
        )

    def _format_date(self, date_value):
        if not date_value:
            return ""
        lang_code = self.env.user.lang or self.env.context.get("lang") or "es_CL"
        lang = self.env["res.lang"]._lang_get(lang_code)
        date_format = lang.date_format if lang else "%d/%m/%Y"
        return date_value.strftime(date_format)

    def _format_datetime(self, value):
        return self._report_service()._format_datetime(value)

    def _ensure_daily_closures(self, users):
        self.ensure_one()
        self.env["surpay.cash.closure"].sudo()._prepare_daily_closures(
            day_value=self.report_date,
            seller_user_ids=users.ids,
        )

    def _get_report_data(self):
        self.ensure_one()
        users = self._get_effective_users()
        self._ensure_daily_closures(users)
        sections = self._build_sections_sql(users)
        company = self.env.company
        logo_b64 = company.logo.decode("utf-8") if company.logo else None
        return {
            "report_date": fields.Date.to_string(self.report_date),
            "report_date_display": self._format_date(self.report_date),
            "generated_at": self._report_service()._format_datetime(fields.Datetime.now()),
            "generated_by": self.env.user.name,
            "company_name": company.name,
            "company_logo": logo_b64,
            "sections": sections,
            "has_data": any(section["has_data"] for section in sections),
        }

    def action_print_report(self):
        self.ensure_one()
        report_data = self._get_report_data()
        if not report_data["has_data"]:
            raise UserError(_("No se encontraron cierres de caja para la fecha y usuarios seleccionados."))
        action = self.env.ref("surpay_base.action_report_surpay_cash_closure").report_action(self)
        action["close_on_report_download"] = True
        return action

    def action_export_excel(self):
        self.ensure_one()
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            raise UserError(_("La librería openpyxl no está instalada en el servidor."))

        report_data = self._get_report_data()
        if not report_data["has_data"]:
            raise UserError(_("No se encontraron cierres de caja para la fecha y usuarios seleccionados."))

        BLUE_FILL = PatternFill("solid", fgColor="1F4E79")
        LIGHTBLUE_FILL = PatternFill("solid", fgColor="DBE5F1")
        HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
        SUBHEADER_FONT = Font(bold=True, size=10)
        BOLD = Font(bold=True)
        CENTER = Alignment(horizontal="center")
        THIN = Side(style="thin", color="BBBBBB")
        CELL_BORDER = Border(left=THIN, right=THIN, bottom=THIN)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        for section in report_data["sections"]:
            sheet_name = section["user"].name[:31]
            ws = wb.create_sheet(title=sheet_name)
            ws.column_dimensions["A"].width = 32
            ws.column_dimensions["B"].width = 26
            ws.column_dimensions["C"].width = 16
            ws.column_dimensions["D"].width = 14
            ws.column_dimensions["E"].width = 30
            ws.column_dimensions["F"].width = 48

            row = 1
            # ---- Encabezado empresa ----
            ws.merge_cells(f"A{row}:E{row}")
            cell = ws[f"A{row}"]
            cell.value = report_data["company_name"]
            cell.font = Font(bold=True, size=14, color="1F4E79")
            row += 1

            ws.merge_cells(f"A{row}:E{row}")
            cell = ws[f"A{row}"]
            cell.value = f"Conciliación Diaria - {report_data['report_date_display']}"
            cell.font = Font(bold=True, size=12)
            row += 2

            # ---- Metadatos ----
            meta = [
                ("Usuario:", section["user"].name),
                ("Fecha de cierre:", report_data["report_date_display"]),
                ("Cierres:", section["closure_names"] or "Sin cierres"),
                ("Estados:", section["closure_states"]),
                ("Generado por:", report_data["generated_by"]),
                ("Generado el:", report_data["generated_at"]),
            ]
            for label, value in meta:
                ws[f"A{row}"] = label
                ws[f"A{row}"].font = BOLD
                ws[f"B{row}"] = value
                row += 1
            row += 1

            # ---- Resumen ----
            ws.merge_cells(f"A{row}:B{row}")
            cell = ws[f"A{row}"]
            cell.value = "Resumen del día"
            cell.font = HEADER_FONT
            cell.fill = BLUE_FILL
            cell.alignment = CENTER
            row += 1

            summary_rows = [
                ("Cierres encontrados", section["closure_count"]),
                ("Transacciones aprobadas", section["transaction_count"]),
                ("Monto bruto procesado", section["gross_amount_display"]),
                ("Comisión Surpay", f"{section['commission_amount_display']} ({section['commission_percent_display']})"),
                ("Monto neto a liquidar", section["net_amount_display"]),
            ]
            for label, value in summary_rows:
                ws[f"A{row}"] = label
                ws[f"B{row}"] = str(value)
                ws[f"A{row}"].border = CELL_BORDER
                ws[f"B{row}"].border = CELL_BORDER
                row += 1
            row += 1

            # ---- Detalle de transacciones ----
            headers = ["Orden", "Fecha / Hora", "Monto", "Estado", "Concepto", "Data extra"]
            cols = ["A", "B", "C", "D", "E", "F"]
            for col, h in zip(cols, headers):
                cell = ws[f"{col}{row}"]
                cell.value = h
                cell.font = SUBHEADER_FONT
                cell.fill = LIGHTBLUE_FILL
                cell.border = CELL_BORDER
            row += 1

            for line in section["transactions"]:
                values = [
                    line["order_id"],
                    line["datetime"],
                    line["amount_display"],
                    line["state_label"],
                    line["concept"],
                    line["extra_data_excel"],
                ]
                for col, val in zip(cols, values):
                    cell = ws[f"{col}{row}"]
                    cell.value = val
                    cell.border = CELL_BORDER
                row += 1
            row += 1

            # ---- Conciliación financiera ----
            ws.merge_cells(f"A{row}:B{row}")
            cell = ws[f"A{row}"]
            cell.value = "Conciliación financiera"
            cell.font = HEADER_FONT
            cell.fill = BLUE_FILL
            cell.alignment = CENTER
            row += 1

            fin_rows = [
                ("Total aprobado (ventas)", section["gross_amount_display"]),
                ("(-) Montos de comisión", section["commission_amount_display"]),
                ("(=) Total conciliado", section["financial_total_display"]),
            ]
            for label, value in fin_rows:
                ws[f"A{row}"] = label
                ws[f"A{row}"].font = BOLD
                ws[f"B{row}"] = value
                ws[f"A{row}"].border = CELL_BORDER
                ws[f"B{row}"].border = CELL_BORDER
                row += 1

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        filename = "Conciliacion_Diaria_{}.xlsx".format(report_data["report_date"])
        attachment = self.env["ir.attachment"].sudo().create({
            "name": filename,
            "type": "binary",
            "datas": base64.b64encode(output.read()).decode("utf-8"),
            "res_model": self._name,
            "res_id": self.id,
            "mimetype": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        })
        return {
            "type": "ir.actions.act_url",
            "url": "/web/content/{}?download=true".format(attachment.id),
            "target": "new",
        }
